import os
import pandas as pd

import tkinter as tk
from tkinter import ttk, filedialog

from unidecode import unidecode

from uuid import uuid4
from openpyxl import load_workbook

from PIL import Image, ImageDraw, ImageTk

global progress_bar, root, cols_organizer, cols_dropped, \
    root_directory, xlsx_files, checkboxes, cols_kept, cols_selected

checkboxes = {}
root_directory = os.getcwd()
progress_bar, root = None, None
cols_organizer, cols_dropped, cols_kept, cols_selected = [], [], [], []

xlsx_files = [
    file for file in os.listdir(root_directory)
    if not file.startswith('~$') and (
        file.endswith(".xlsx") or file.endswith(".xls"))
    ]

enc = 'utf-8'


def create_image():
    image = Image.new('RGB', (18, 18), color='white')
    draw = ImageDraw.Draw(image)
    draw.text((0, -1), 'X', fill='black')
    draw.text((6, -1), 'C', fill='blue')
    draw.text((12, -1), 'L', fill='green')
    draw.text((0, 8), 'C', fill='black')
    draw.text((6, 8), 'C', fill='blue')
    draw.text((12, 8), 'T', fill='green')
    return ImageTk.PhotoImage(image)


def organize_col(x):
    for i in cols_organizer:
        if i[0] == 's':
            if (
                    x.replace("'", '').strip().strip(
                        '_').upper()).startswith(i[1]):
                return i[2]
        elif i[0] == '=':
            if (
                    x.replace("'", '').strip().strip(
                        '_').upper()) == i[1]:
                return i[2]
    return (
        x.replace("'", '').strip().strip('_').upper())


def get_closest_col(dat, length):
    closest_column = None
    closest_column_index = None
    min_diff = float('inf')
    for i, column in enumerate(dat.columns):
        numeric_strings = dat[column].apply(
            lambda x: str(x) if str(x).isdigit() else None)
        avg_length = numeric_strings.dropna().apply(len).mean()
        diff = abs(avg_length - length)
        if diff < min_diff:
            closest_column = column
            closest_column_index = i
            min_diff = diff
    return closest_column, closest_column_index


def find_header_row(df, closest_col):
    for i in range(df.shape[0]):
        try:
            if df.iloc[i].astype(str).str.match(
                    '^\d{5}$').get(closest_col, False):
                return i           
        except:
            continue
    return None


def process_df(df, length):
    closest_col, closest_col_index = get_closest_col(df, length)
    header_row = find_header_row(df, closest_col)
    if header_row is None:
        return
    final_col_names = ["" for _ in range(len(df.columns))]
    for row in range(0, header_row):
        for i, col in enumerate(df.columns):
            if ('unnamed' not in str(df.iloc[row, i]).lower()
                    and str(df.iloc[row, i]).lower() != 'nan'
                    and str(df.iloc[row, i]).strip() != ""):
                if final_col_names[i]:
                    final_col_names[i] += " " + str(
                        df.iloc[row, i]).strip()
                else:
                    final_col_names[i] = str(
                        df.iloc[row, i]).strip()
    df.columns = [(col.replace("'", '').strip().strip(
        '_').upper().replace(' ', '_'))
        for col in final_col_names]
    df = df[header_row:].reset_index(drop=True)
    return df[df.iloc[:, closest_col_index].apply(
        lambda x: len(str(x)) == length and
        str(x).isdigit())], df.columns[closest_col_index]


def remove_duplicate_columns(df):
    return df.loc[:, ~df.columns.duplicated()]


def rename_and_filter_columns(df, cols_kept, closest_col,
                              sheet_name, length):
    df.columns = [
        (col.replace("'", '').strip().strip(
            '_').upper().replace(' ', '_'))
            for col in df.columns]
    cols_to_keep = set()
    for rule in cols_kept:
        parts = rule.split()
        if len(parts) == 3:
            command = parts[0]
            if command == 'len':
                target_col_name = parts[2]
                df.rename(columns={closest_col: target_col_name},
                        inplace=True)
                cols_to_keep.add(target_col_name)
            elif command == '=':
                current_col_name = parts[1]
                new_col_name = parts[2]
                if current_col_name not in df.columns:
                    open_popup(f'col {current_col_name} no existe')
                df.rename(columns={current_col_name: new_col_name},
                        inplace=True)
                cols_to_keep.add(new_col_name)
            elif command == 'c':
                contains_values = parts[1].split(',')
                new_col_name = parts[2]
                for col in df.columns:
                    if all(value in col for value in contains_values):
                        df.rename(columns={col: new_col_name},
                                  inplace=True)
                        break
                cols_to_keep.add(new_col_name)
            elif command == '+':
                col_names = parts[1].split(',')
                new_col_name = parts[2]
                if 'SHEETNAME' in col_names:
                    col_names.remove('SHEETNAME')
                    df[new_col_name] = df[col_names].apply(
                        lambda x: ''.join(
                            x.astype(str)), axis=1) + sheet_name
                else:
                    df[new_col_name] = df[col_names].apply(
                        lambda x: ''.join(x.astype(str)), axis=1)
                cols_to_keep.add(new_col_name)
            elif command == 'z':
                current_col_name = parts[1]
                new_col_name = parts[2]
                df[new_col_name] = df[current_col_name].apply(
                    lambda x: str(x).zfill(length))
                cols_to_keep.add(new_col_name)
        else:
            open_popup(f'Error en la configuración:\n{rule}')
    df = df[[i for i in list(cols_to_keep) if i in  df.columns]]
    return df


def open_popup(message):
    popup = tk.Toplevel()
    popup.title("Message")
    # Set the position of the popup
    popup.geometry("300x100")
    # Add a label with the message
    message_label = tk.Label(popup, text=message, wraplength=280)
    message_label.pack(pady=20)
    # Add a button to close the popup
    close_button = tk.Button(popup, text="Close", command=popup.destroy)
    close_button.pack(pady=5)


def make_excel_concat(selected_files: list):
    global progress_bar, root, cols_organizer, \
        cols_dropped, cols_kept, cols_selected
    load_vars()
    progress_bar.start()    
    concat_filename = str(uuid4())
    if cols_kept:
        column_dict = list(
            set([j for i in cols_kept
                 if ((j := i.split()[-1]) not in cols_dropped) and
                 ((j := i.split()[-1]) in cols_selected if cols_selected else True)]))
        with open(get_filename('_logs_columnas'), 'w', encoding=enc) as cols:
            cols.write(f'Columns found: {len(column_dict)}\n')
            for i in column_dict:
                cols.write(i+'\n')
        print(len(column_dict))
        df_concat = pd.DataFrame()
        # Variable for first iteration headers
        first = True
        for i, filename in enumerate(selected_files):
            if filename.endswith(
                    '.xlsx') or filename.endswith('.xls'):
                print(filename)
                xl_file = pd.ExcelFile(filename)
                # Loop over all sheets in the Excel file
                for sheet_name in xl_file.sheet_names:
                    print(sheet_name)
                    # Read the sheet data and re-order the columns
                    df = pd.read_excel(xl_file, sheet_name=sheet_name,
                                       dtype=str, header=None)
                    length = [i.split()[1]
                              for i in cols_kept
                              if 'len ' in i][0]
                    if length.isnumeric():
                        length = int(length)
                    else:
                        open_popup(f'len {length}, debe ser un numero')
                    df, closest_col = process_df(df, length)
                    df = remove_duplicate_columns(df)
                    df = rename_and_filter_columns(df, cols_kept,
                                                   closest_col,
                                                   sheet_name,
                                                   length)
                    missing_columns = set(column_dict
                                          ) - set(df.columns)
                    for column in missing_columns:
                        df[column] = ''
                    df = df[column_dict]
                    print(df.shape)
                    df = df.reset_index(drop=True)
                    df_concat = pd.concat([df_concat, df], axis=0)
                    root.update()
                df_concat.to_csv(f'{concat_filename}.csv', index=False,
                                 mode='a', header=True if first else False,
                                 sep=';', encoding=enc, errors='ignore')
                first = False
                df_concat = pd.DataFrame()
                print(f'added {filename}')
            root.update()
            progress_bar["value"] = int(i*len(selected_files)/
                                        (3*len(selected_files))) + 33
            root.update()
    else:
        # Initialize an empty dictionary to hold the column names
        column_dict = {}
        # Loop over all Excel files in the current directory
        for i, filename in enumerate(selected_files):
            if filename.endswith('.xlsx'):
                print(filename)
                # Load the Excel workbook
                wb = load_workbook(filename=filename, read_only=True,
                                   data_only=True)
                # Loop over all sheets in the Excel file
                for sheet_name in wb.sheetnames:
                    if sheet_name != 'skip':
                        print(sheet_name)
                        # Read the first row of the sheet to get the column names
                        sheet = wb[sheet_name]
                        columns = [
                            (col.value.replace("'", '').strip().strip(
                                '_').upper().replace(' ', '_'))
                                for col in next(sheet.iter_rows()) if col.value]
                        # Add the cleaned column names to the dictionary of sets
                        if sheet_name in column_dict:
                            column_dict[sheet_name].update(set(columns))
                        else:
                            column_dict[sheet_name] = set(columns)
                        root.update()
            root.update()
            progress_bar["value"] = int(i*len(selected_files)/
                                        (3*len(selected_files)))
            root.update()
        # Turn the sets into list to keep the order
        temp_set = set()
        for i in list(set(column_dict)):
            temp_set = temp_set | column_dict[i]
        column_dict = []
        for x in sorted(list(set([organize_col(x)
                                  for x in list(temp_set)]))):
            column_dict.append(x)
        column_dict = [
            i for i in column_dict
            if (i not in cols_dropped) and (i in cols_selected if cols_selected else True)]
        with open(get_filename('_logs_columnas'), 'w', encoding=enc) as cols:
            cols.write(f'Columns found: {len(column_dict)}\n')
            for i in column_dict:
                cols.write(i+'\n')
        print(len(column_dict))
        df_concat = pd.DataFrame()
        # Variable for first iteration headers
        first = True
        # Loop over all Excel files in the current directory
        for i, filename in enumerate(selected_files):
            if filename.endswith(
                    '.xlsx') or filename.endswith('.xls'):
                print(filename)
                xl_file = pd.ExcelFile(filename)
                # Loop over all sheets in the Excel file
                for sheet_name in xl_file.sheet_names:
                    print(sheet_name)
                    # Read the sheet data and re-order the columns
                    df = pd.read_excel(xl_file,
                                       sheet_name=sheet_name,
                                       dtype=str)         
                    df.rename(columns=lambda x: organize_col(x),
                              inplace=True)
                    missing_columns = set(column_dict
                                          ) - set(df.columns)
                    for column in missing_columns:
                        df[column] = ''
                    df = df[column_dict]
                    print(df.shape)
                    df = df.reset_index(drop=True)
                    df_concat = pd.concat([df_concat, df], axis=0)
                    root.update()
                # Write the concatenated DataFrame to a CSV file
                df_concat.to_csv(f'{concat_filename}.csv',
                                 index=False, mode='a',
                                 header=True if first else False,
                                 sep=';', encoding=enc, errors='ignore')
                first = False
                df_concat = pd.DataFrame()
                print(f'added {filename}')
            root.update()
            progress_bar["value"] = int(i*len(selected_files)/
                                        (3*len(selected_files))) + 33
            root.update()
    return concat_filename


def delete_duplicates(filename: str):
    print('deleting duplicates')
    df = pd.read_csv(f'{filename}.csv', sep=';', dtype='str',
                     encoding=enc, on_bad_lines='skip',
                     encoding_errors='ignore')
    print('read the csv file')
    df = df.map(lambda x: x.encode(enc, 'replace').decode(enc)
                                   if isinstance(x, str) else x)
    # Check for duplicates and save them to Excel if they exist
    if duplicados_var.get():
        if (duplicates := df.duplicated()).any():
            df[duplicates].to_excel(f'duplicados_{filename}.xlsx',
                                    index=False)
        # Drop duplicates based on all columns
        df = df.drop_duplicates()
        print('duplicates deleted')
    # Write the DataFrame back to the CSV file
    print('writing to new files')
    if excel_var.get() and csv_var.get():
        df.to_csv(f'{filename}.csv',
                index=False, sep=';',
                encoding=enc, errors='ignore')
        print('wrote to new csv')
        df.to_excel(f'{filename}.xlsx', index=False,
                    engine='openpyxl')
        print('wrote to new excel')
    elif csv_var.get() and not excel_var.get():
        df.to_csv(f'{filename}.csv',
                index=False, sep=';',
                encoding=enc, errors='ignore')
        print('wrote to new csv')
    elif excel_var.get() and not csv_var.get():
        df.to_excel(f'{filename}.xlsx', index=False,
                    engine='openpyxl')
        print('wrote to new excel')
        os.remove(f'{filename}.csv')
    print('finished')


def get_filename(str_):
        return os.path.join('config_concatenador',
                            f'{str_}.txt')

def load_vars():
    global cols_organizer, cols_dropped, cols_kept, cols_selected
    if not os.path.exists('config_concatenador'):
        os.makedirs('config_concatenador')
    
    def make_file(file_name):
        if not os.path.isfile(file_name):
            with open(file_name, 'w', encoding=enc) as file:
                file.write("")

    if renombrar_var.get():
        file_name = get_filename('cols_renombrar')
        try:
            with open(file_name, 'r', encoding=enc) as f:
                cols_organizer = [
                    i.strip().split()
                    for i in f.readlines() if i]
        except:
            make_file(file_name)
            cols_organizer = []
    if quitar_var.get():
        file_name = get_filename('cols_quitar')
        try:
            with open(file_name, 'r', encoding=enc) as f:
                cols_dropped = [
                    i.strip()
                    for i in f.readlines() if i]
        except:
            make_file(file_name)
            cols_dropped = []
    if seleccion_var.get():
        file_name = get_filename('cols_seleccion')
        try:
            with open(file_name, 'r', encoding=enc) as f:
                cols_selected = [
                    i.strip()
                    for i in f.readlines() if i]
        except:
            make_file(file_name)
            cols_selected = []
    if poblacion_var.get():
        file_name = get_filename('cols_poblacion_riesgo')
        try:
            with open(file_name, 'r', encoding=enc) as f:
                cols_kept = [
                    i.strip()
                    for i in f.readlines() if i]
        except:
            make_file(file_name)
            cols_kept = []


def get_selected_files():
    global progress_bar
    progress_bar = ttk.Progressbar(root,
                                   orient='horizontal',
                                   length=350,
                                   mode="determinate")
    progress_bar.pack(pady=20)
    progress_bar["value"] = 0
    selected_files = []
    for filename, var in checkboxes.items():
        if var.get():
            selected_files.append(filename)
    if selected_files:
        filename = make_excel_concat(selected_files)
        progress_bar["value"] = 75
        delete_duplicates(filename)
        progress_bar["value"] = 100
    progress_bar.destroy()


def get_all_files():
    global progress_bar
    progress_bar = ttk.Progressbar(root,
                                   orient='horizontal',
                                   length=350,
                                   mode="determinate")
    progress_bar.pack(pady=20)
    progress_bar["value"] = 0
    selected_files = []
    for filename, var in checkboxes.items():
        selected_files.append(filename)
    filename = make_excel_concat(selected_files)
    progress_bar["value"] = 75
    delete_duplicates(filename)
    progress_bar["value"] = 100
    progress_bar.destroy()


def set_current_directory():
    global root_directory
    root_directory = os.getcwd()
    update_file_label()
    update_excel_files()


def browse_directory():
    global root_directory
    selected_directory = filedialog.askdirectory()
    if selected_directory:
        root_directory = selected_directory
    else:
        root_directory = os.getcwd()
    update_file_label()
    update_excel_files()
    root.geometry(f"450x{25*len(xlsx_files) + 375}")


def update_file_label():
    max_line_length = 60
    display_text = f"Archivos en la carpeta: {root_directory}"
    # Wrap the text if it's too long
    if len(display_text) > max_line_length:
        space_index = display_text.rfind(' ', 0, max_line_length)
        if space_index != -1:
            display_text = display_text[:space_index
                                        ] + '\n' + display_text[
                                            space_index+1:]
    file_label.config(text=display_text)


def update_excel_files():
    global xlsx_files, checkboxes, checkboxes_frame
    # Clear existing checkboxes
    for widget in checkboxes_frame.winfo_children():
        widget.destroy()
    checkboxes = {}
    # Get new list of Excel files
    xlsx_files = [
        os.path.join(root_directory, file)
        for file in os.listdir(root_directory)
        if not file.startswith('~$') and
        (file.endswith(".xlsx") or file.endswith(".xls"))]
    # Create a checkbox for each new .xlsx file
    for file in xlsx_files:
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(
            checkboxes_frame,
            text=os.path.basename(file),
            variable=var)
        checkbox.pack(anchor=tk.W)
        checkboxes[file] = var


if __name__ == '__main__':
    root = tk.Tk()
    root.title("Concatenador de Excel")
    icon_image = create_image()
    # Set the image as the app's icon
    root.iconphoto(True, icon_image)
    # Get all files with the .xlsx extension in the root directory
    root.geometry(f"450x{25*len(xlsx_files) + 375}")
    title_label = tk.Label(
        root, text="Seleccione los archivos para concatenar",
        font=("Arial", 14))
    title_label.pack(pady=(5, 10))
    separator = ttk.Separator(root, orient='horizontal')
    separator.pack(fill='x', padx=5, pady=5)
    cwd_frame = tk.Frame(root)
    cwd_frame.pack()
    # Button to set current directory
    current_dir_button = tk.Button(cwd_frame,
                                   text="Directorio actual",
                                   command=set_current_directory)
    current_dir_button.pack(side=tk.LEFT, padx=10, pady=5)
    # Button to browse for a directory
    browse_dir_button = tk.Button(cwd_frame,
                                  text="Buscar directorio",
                                  command=browse_directory)
    browse_dir_button.pack(side=tk.LEFT, padx=10, pady=5)
    file_label = tk.Label(
        root, text=f"Archivos en la carpeta: {root_directory}",
        font=("Arial", 10))
    file_label.pack(pady=(5, 10))
    separator = ttk.Separator(root, orient='horizontal')
    separator.pack(fill='x', padx=5, pady=5)
    # Frame for checkboxes
    checkboxes_frame = tk.Frame(root)
    checkboxes_frame.pack()
    update_excel_files()
    spacer_frame = tk.Frame(root)
    spacer_frame.pack()
    # Frame for "Concatenar" buttons
    concatenar_frame = tk.Frame(spacer_frame)
    concatenar_frame.pack(side=tk.LEFT)
    # Spacer
    spacer = tk.Frame(concatenar_frame, height=20)
    spacer.pack()
    # Concatenate selected files button
    select_button = tk.Button(concatenar_frame,
                              text="Concatenar Seleccionados",
                              command=get_selected_files)
    select_button.pack()
    spacer1 = tk.Frame(concatenar_frame, height=10)
    spacer1.pack()
    select_all_button = tk.Button(concatenar_frame,
                                  text="Concatenar Todos",
                                  command=get_all_files)
    select_all_button.pack()
    format_frame = tk.Frame(spacer_frame)
    format_frame.pack(side=tk.LEFT, padx=30, pady=20)
    excel_var = tk.BooleanVar(value=True, master=root)
    csv_var = tk.BooleanVar(value=False, master=root)
    extension_label = tk.Label(format_frame, text='Extensión:')
    extension_label.pack(side=tk.TOP)
    excel_checkbox = tk.Checkbutton(format_frame,
                                  text='XLSX',
                                  variable=excel_var)
    excel_checkbox.pack(side=tk.TOP)
    csv_checkbox = tk.Checkbutton(format_frame,
                                  text='CSV',
                                  variable=csv_var)
    csv_checkbox.pack(side=tk.TOP)
    checkbox_frame = tk.Frame(spacer_frame)
    checkbox_frame.pack(pady=10)
    # Variables to track checkbox states
    poblacion_var = tk.BooleanVar(value=False, master=root)
    quitar_var = tk.BooleanVar(value=False, master=root)
    seleccion_var = tk.BooleanVar(value=False, master=root)
    renombrar_var = tk.BooleanVar(value=False, master=root)
    duplicados_var = tk.BooleanVar(value=False, master=root)
    extension_label = tk.Label(checkbox_frame, text='Configuración:')
    extension_label.pack(side=tk.TOP)
    # Checkboxes
    poblacion_checkbox = tk.Checkbutton(checkbox_frame,
                                        text='Poblacion Riesgo',
                                        variable=poblacion_var)
    poblacion_checkbox.pack(side=tk.TOP, anchor='w')
    quitar_checkbox = tk.Checkbutton(checkbox_frame,
                                     text='Quitar Columnas',
                                     variable=quitar_var)
    quitar_checkbox.pack(side=tk.TOP, anchor='w')
    seleccion_checkbox = tk.Checkbutton(checkbox_frame,
                                     text='Seleccionar Columnas',
                                     variable=seleccion_var)
    seleccion_checkbox.pack(side=tk.TOP, anchor='w')
    renombrar_checkbox = tk.Checkbutton(checkbox_frame,
                                        text='Renombrar Columnas',
                                        variable=renombrar_var)
    renombrar_checkbox.pack(side=tk.TOP, anchor='w')
    duplicados_checkbox = tk.Checkbutton(checkbox_frame,
                                        text='Borrar Duplicados',
                                        variable=duplicados_var)
    duplicados_checkbox.pack(side=tk.TOP, anchor='w')
    root.mainloop()
